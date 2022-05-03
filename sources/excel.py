#!/bin/python3

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from sources.classes import User

HEADER_REF: list[str] = ["Classement", "GPA", "Prénom", "Nom", "Login", "Ville"]
HEADER_CITY_REF: list[str] = ["Classement", "Moyenne GPA", "Ville", "Nombre d'étudiants"]


def sort_cities(cities: dict) -> list:
    dest: list = []
    for city in cities:
        moyenne = 0
        for user in cities[city]:
            moyenne += float(user.gpa)
        moyenne /= len(cities[city])
        dest.append((moyenne, city, len(cities[city])))
    dest.sort(key=lambda elem: elem[0], reverse=True)
    return dest


class Excel:
    def __init__(self):
        self.workbook: Workbook = Workbook()

    def save(self):
        self.workbook.save("./data/results.xlsx")

    @staticmethod
    def add_users_to_worksheet(users: list[User], work_sheet: Worksheet):
        header = work_sheet["A1":"F1"]
        for index, value in enumerate(HEADER_REF):
            header[0][index].value = value
        table = work_sheet["A2":"F%d" % (len(users) + 1)]
        for index, user in enumerate(users):
            table[index][0].value = index + 1
            table[index][1].value = user.gpa
            table[index][2].value = user.firstname
            table[index][3].value = user.lastname
            table[index][4].value = user.login
            table[index][5].value = user.city

    def add_all_users(self, users: list[User]):
        work_sheet: Worksheet = self.workbook.active
        work_sheet.title = "All"
        self.add_users_to_worksheet(users, work_sheet)

    def add_cities(self, cities: dict):
        for city in cities:
            work_sheet: Worksheet = self.workbook.create_sheet(city)
            self.add_users_to_worksheet(cities[city], work_sheet)

    def cities_rank(self, cities: dict):
        work_sheet: Worksheet = self.workbook.create_sheet("Classement par ville")
        sorted_cities = sort_cities(cities)
        header = work_sheet["A1":"D1"]
        for index, value in enumerate(HEADER_CITY_REF):
            header[0][index].value = value
        table = work_sheet["A2":"F%d" % (len(sorted_cities) + 1)]
        for index, city in enumerate(sorted_cities):
            table[index][0].value = index + 1
            table[index][1].value = city[0]
            table[index][2].value = city[1]
            table[index][3].value = city[2]


EXCEL: Excel = Excel()
